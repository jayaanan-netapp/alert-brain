import mysql.connector
from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.svm import LinearSVC
from sklearn.model_selection import train_test_split
from sklearn import metrics
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.decomposition import NMF
import pandas as pd
import tensorflow as tf
from sklearn.model_selection import train_test_split
from transformers import TFBertModel, BertTokenizer
from sklearn.metrics import classification_report, confusion_matrix

import re

import openpyxl
from transformers import Trainer, TrainingArguments, DistilBertForSequenceClassification, DistilBertTokenizerFast, \
     DataCollatorWithPadding, pipeline
from datasets import Dataset, metric
import numpy as np
import evaluate

'''CREATE TABLE `alerts` (
	`id` INT(10) NOT NULL AUTO_INCREMENT,
	`message` MEDIUMTEXT NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`gpt_category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`svm_category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`nb_category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`lda_category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`nmf_category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	`bert_category` VARCHAR(1000) NULL DEFAULT NULL COLLATE 'utf8mb4_0900_ai_ci',
	INDEX `Index 1` (`id`) USING BTREE
)
COLLATE='utf8mb4_0900_ai_ci'
ENGINE=InnoDB
;
'''

# Modify the display options
pd.set_option('display.max_columns', None)  # Show all columns
pd.set_option('display.max_rows', None)  # Show all rows
pd.set_option('display.width', None)  # Auto-adjust the display width


class AlertBrain:

    def __init__(self, host, user, password, database):
        self.df = pd.DataFrame(columns=['category', 'message'], dtype=object)
        self.conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=database)
        self.conn.connect()
        self.cursor = self.conn.cursor()
        self.tokenizer = DistilBertTokenizerFast.from_pretrained('distilbert-base-cased')
        self.metric = evaluate.load("accuracy")

    def close(self):
        self.cursor.close()
        if self.conn.is_connected():
            self.conn.close()
            print("Connection closed")

    def compute_metrics(self, eval_pred):  # custom method to take in logits and calculate accuracy of the eval set
        logits, labels = eval_pred
        predictions = np.argmax(logits, axis=-1)
        return self.metric.compute(predictions=predictions, references=labels)

    # simple function to batch tokenize utterances with truncation
    def preprocess_function(self, examples):
        return self.tokenizer(examples["utterance"], truncation=True)

    def load_and_clean_data(self, file_path, sheet_name):
        # Open the Excel file
        wb = openpyxl.load_workbook(file_path)
        # Get the second sheet
        sheet = wb[sheet_name]
        # Iterate over rows in the sheet
        for row in sheet.iter_rows(values_only=True):
            # Convert any None values to NULL
            row = [value if value is not None else 'NULL' for value in row]
            alerts = row[1].splitlines()
            # print(alerts)
            # Using a for loop
            for string in alerts:
                alert = re.sub(r"\b\d+\.", "", string, count=1)
                # print(row[0], alert)
                query = "INSERT INTO alerts (category, message) VALUES (%s, %s)"
                values = (row[0], alert)
                self.df = self.df.append({'category': row[0], 'message': alert}, ignore_index=True)

                # Executing the query with the values
                self.cursor.execute(query, values)
                self.conn.commit()
        # print(df)
        # Check for the existence of NaN values in a cell:
        self.df.isnull().sum()
        self.df.dropna(inplace=True)
        # Delete and remove empty strings.
        blanks = []  # start with an empty list
        for i, lb, rv in self.df.itertuples():  # iterate over the DataFrame
            if type(rv) == str:  # avoid NaN values
                if rv.isspace():  # test 'review' for whitespace
                    blanks.append(i)  # add matching index numbers to the list
        # print(len(blanks), 'blanks: ', blanks)
        self.df.drop(blanks, inplace=True)

    def execute_supervised_learning_modal(self, text_clf_lsvc, column_name):

        X = self.df['message']
        y = self.df['category']
        '''
        Section for supervised learning. Hear we are evaluating NB and SVM to train and test on alerts.
        '''
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.33, random_state=42)

        text_clf_lsvc.fit(X_train, y_train)
        # Form a prediction set
        predictions = text_clf_lsvc.predict(X_test)
        # Get corresponding texts for the predictions
        predicted_texts = []
        for prediction, text in zip(predictions, X_test):
            predicted_texts.append((prediction, text))
        # Print the predictions and corresponding texts
        for prediction, text in predicted_texts:
            '''print("Prediction:", prediction)
            print("Text:", text)
            print()'''
            query = 'SELECT * FROM alerts WHERE message = %s'
            self.cursor.execute(query, (text,))
            rows = self.cursor.fetchall()
            for row in rows:
                query = f"UPDATE alerts SET {column_name} = %s WHERE id = %s"
                self.cursor.execute(query, (prediction, row[0]))
        # Update test data.
        for train in X_train:
            query = 'SELECT * FROM alerts WHERE message = %s'
            self.cursor.execute(query, (train,))
            rows = self.cursor.fetchall()
            for row in rows:
                query = f"UPDATE alerts SET {column_name} = %s WHERE id = %s"
                self.cursor.execute(query, (row[2], row[0]))
        # Commit the changes to the database
        print(metrics.confusion_matrix(y_test, predictions))
        # Print a classification report
        print(metrics.classification_report(y_test, predictions))
        # Print the overall accuracy
        print(metrics.accuracy_score(y_test, predictions))
        self.conn.commit()

    def un_supervise_learning(self):
        cv = CountVectorizer(max_df=0.9, min_df=2, stop_words='english')
        dtm = cv.fit_transform(self.df['message'])
        LDA = LatentDirichletAllocation(n_components=5, random_state=42)
        LDA.fit(dtm)
        for index, topic in enumerate(LDA.components_):
            print(f'THE TOP 3 WORDS FOR TOPIC #{index}')
            print([cv.get_feature_names()[i] for i in topic.argsort()[-3:]])
        topic_results = LDA.transform(dtm)
        self.df["LDA_Topic"] = topic_results.argmax(axis=1)
        lda_topic_dictionary = {0: 'memory', 1: 'virtual_machine', 2: 'performance', 3: 'connection', 4: 'cloud'}
        self.df["LDA_Topic_label"] = self.df["LDA_Topic"].map(lda_topic_dictionary)
        print(self.df)

        tfidf = TfidfVectorizer(max_df=0.9, min_df=2, stop_words='english')
        dtm = tfidf.fit_transform(self.df['message'])
        nmf_model = NMF(n_components=5, random_state=42)
        nmf_model.fit(dtm)
        for index, topic in enumerate(nmf_model.components_):
            print(f'THE TOP 3 WORDS FOR TOPIC #{index}')
            print([tfidf.get_feature_names()[i] for i in topic.argsort()[-3:]])
        topic_results = LDA.transform(dtm)
        self.df["NMF_Topic"] = topic_results.argmax(axis=1)
        lda_topic_dictionary = {0: 'service_alert', 1: 'hypervisor', 2: 'cloud', 3: 'data_inconsistent',
                                4: 'kubernetes'}
        self.df["NMF_Topic_label"] = self.df["NMF_Topic"].map(lda_topic_dictionary)
        print(self.df)
        # Iterate over the DataFrame rows and insert data related to unsupervised learning
        for row in self.df.iterrows():
            # Get the text from the row
            text = row[1]['message']
            query = 'SELECT * FROM alerts WHERE message = %s'
            self.cursor.execute(query, (text,))
            rows = self.cursor.fetchall()
            for data_row in rows:
                query = 'UPDATE alerts SET lda_category = %s WHERE id = %s'
                self.cursor.execute(query, (row[1]['LDA_Topic_label'], data_row[0]))
                query = 'UPDATE alerts SET nmf_category = %s WHERE id = %s'
                self.cursor.execute(query, (row[1]['NMF_Topic_label'], data_row[0]))
        self.conn.commit()

    def supervise_learning(self):
        # Naïve Bayes:
        text_clf_nb = Pipeline([('tfidf', TfidfVectorizer()),
                                ('clf', MultinomialNB()),
                                ])
        self.execute_supervised_learning_modal(text_clf_nb, 'nb_category')

        # Linear SVC:
        text_clf_lsvc = Pipeline([('tfidf', TfidfVectorizer()),
                                  ('clf', LinearSVC()),
                                  ])
        self.execute_supervised_learning_modal(text_clf_lsvc, 'svm_category')

    def deep_learning(self):
        wb = openpyxl.load_workbook('C:/Users/jayaanan/Downloads/Alerts_Data1.xlsx')
        # Get the second sheet
        sheet = wb['Sheet2']
        utterances = []
        sequence_labels = []

        for row in sheet.iter_rows(values_only=True):
            # Convert any None values to NULL
            row = [value if value is not None else 'NULL' for value in row]
            alerts = row[1].splitlines()
            # print(alerts)
            # Using a for loop
            for string in alerts:
                alert = re.sub(r"\b\d+\.", "", string, count=1)
                utterances.append(alert)
                sequence_labels.append(row[0])

        print('length of all tokens', len(utterances), len(sequence_labels))
        print('first utterance', utterances[0], 'sequence label', sequence_labels[0])
        tokenizer = DistilBertTokenizerFast.from_pretrained('distilbert-base-cased')
        unique_sequence_labels = list(set(sequence_labels))
        print('unique_sequence_labels-->', unique_sequence_labels)
        sequence_labels = [unique_sequence_labels.index(l) for l in sequence_labels]
        print(f'There are {len(unique_sequence_labels)} unique sequence labels')

        print('utterances', utterances[0], len(utterances))
        print('sequence_labels', sequence_labels[0], len(sequence_labels))
        print(unique_sequence_labels[sequence_labels[0]])

        snips_dataset = Dataset.from_dict(
            dict(
                utterance=utterances,
                label=sequence_labels
            )
        )

        snips_dataset = snips_dataset.train_test_split(test_size=0.2)
        print(snips_dataset)

        #print(tokenizer('hi'))
        #print(tokenizer.decode([101, 2603, 1142, 18977, 126, 2940, 102]))

        seq_clf_tokenized_snips = snips_dataset.map(self.preprocess_function, batched=True)

        # only input_ids, attention_mask, and label are used. The rest are for show
        #print(seq_clf_tokenized_snips['train'][0])

        # DataCollatorWithPadding creates batch of data. It also dynamically pads text to the
        #  length of the longest element in the batch, making them all the same length.
        #  It's possible to pad your text in the tokenizer function with padding=True, dynamic padding is more efficient.

        data_collator = DataCollatorWithPadding(tokenizer=tokenizer)
        #print({i: l for i, l in enumerate(unique_sequence_labels)})

        sequence_clf_model = DistilBertForSequenceClassification.from_pretrained(
            'distilbert-base-cased',
            num_labels=len(unique_sequence_labels),
        )

        # set an index -> label dictionary
        sequence_clf_model.config.id2label = {i: l for i, l in enumerate(unique_sequence_labels)}
        print(sequence_clf_model.config.id2label[0])

        epochs = 5

        training_args = TrainingArguments(
            output_dir="./alert_brain/results",
            num_train_epochs=epochs,
            per_device_train_batch_size=32,
            per_device_eval_batch_size=32,
            load_best_model_at_end=True,

            # some deep learning parameters that the Trainer is able to take in
            warmup_steps=len(seq_clf_tokenized_snips['train']) // 5,
            # number of warmup steps for learning rate scheduler,
            weight_decay=0.05,

            logging_steps=1,
            log_level='info',
            evaluation_strategy='epoch',
            eval_steps=50,
            save_strategy='epoch'
        )

        # Define the trainer:

        trainer = Trainer(
            model=sequence_clf_model,
            args=training_args,
            train_dataset=seq_clf_tokenized_snips['train'],
            eval_dataset=seq_clf_tokenized_snips['test'],
            compute_metrics=self.compute_metrics,  # optional
            data_collator=data_collator
        )
        print(trainer.evaluate())
        trainer.train()
        print(trainer.evaluate())
        trainer.save_model()
        # We can now load our fine-tuned from our directory
        pipe = pipeline("text-classification", "./alert_brain/results", tokenizer=tokenizer)

        for index, utterance in enumerate(utterances):
            prediction = pipe(utterance)
            #print(utterance, prediction, sequence_labels[index])
            query = 'SELECT * FROM alerts WHERE message = %s'
            self.cursor.execute(query, (utterance,))
            rows = self.cursor.fetchall()
            for row in rows:
                query = f"UPDATE alerts SET bert_category = %s WHERE id = %s"
                self.cursor.execute(query, (prediction[0]['label'], row[0]))
        self.conn.commit()


if __name__ == '__main__':
    brain = AlertBrain("localhost", "root", "Netapp!2345", "alert_brain")
    brain.load_and_clean_data('C:/Users/jayaanan/Downloads/Alerts_Data1.xlsx', 'Sheet2')
    brain.supervise_learning()
    brain.un_supervise_learning()
    brain.deep_learning()
    brain.close()
